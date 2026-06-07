"""
Export map/data/states.js bill data to Excel.
Run: python export_map.py
Output: map_bills_export.xlsx
"""

import json
import re
from pathlib import Path
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

STATES_JS = Path("map/data/states.js")
OUT_FILE   = Path("map_bills_export.xlsx")


def load_states():
    text = STATES_JS.read_text(encoding="utf-8")
    # Strip JS variable declaration and trailing semicolon
    text = re.sub(r"^.*?const\s+stateData\s*=\s*", "", text, flags=re.DOTALL)
    text = text.strip().rstrip(";").strip()
    # Strip // comments only when NOT inside a string — scan char by char
    out = []
    in_str = False
    escape = False
    i = 0
    while i < len(text):
        ch = text[i]
        if escape:
            out.append(ch)
            escape = False
        elif ch == '\\' and in_str:
            out.append(ch)
            escape = True
        elif ch == '"' :
            in_str = not in_str
            out.append(ch)
        elif not in_str and ch == '/' and i + 1 < len(text) and text[i + 1] == '/':
            # skip to end of line
            while i < len(text) and text[i] != '\n':
                i += 1
            continue
        else:
            out.append(ch)
        i += 1
    text = "".join(out)
    # Quote unquoted JS object keys (identifiers before colon not already quoted)
    text = re.sub(r'(?<!["\w])([A-Za-z_][A-Za-z0-9_]*)(\s*:)', r'"\1"\2', text)
    # Remove trailing commas before } or ]
    text = re.sub(r',(\s*[}\]])', r'\1', text)
    return json.loads(text)


def flatten(state_data: dict) -> list[dict]:
    rows = []
    for abbr, state in state_data.items():
        name    = state.get("name", abbr)
        pol     = state.get("politics") or {}
        gov     = pol.get("governor", "")
        market  = pol.get("market", "")
        chambers = pol.get("chambers", {})

        # Build chamber string e.g. "Senate: R, House: R"
        chamber_str = ", ".join(f"{k}: {v}" for k, v in chambers.items())

        # Trifecta
        parties = [gov] + list(chambers.values())
        if all(p == "R" for p in parties if p):
            trifecta = "R-trifecta"
        elif all(p == "D" for p in parties if p):
            trifecta = "D-trifecta"
        elif parties and any(p for p in parties):
            trifecta = "Divided"
        else:
            trifecta = ""

        fields = state.get("fields", {})
        if not fields:
            rows.append({
                "State": name, "Abbr": abbr,
                "Governor": gov, "Chambers": chamber_str,
                "Trifecta": trifecta, "Market": market,
                "Bill": "", "Summary": "", "URL": "",
            })
            continue

        for bill_no, value in fields.items():
            if isinstance(value, dict):
                text = value.get("text", "")
                url  = value.get("url", "")
            elif isinstance(value, str):
                text = value
                url  = ""
            else:
                text = url = ""

            rows.append({
                "State":    name,
                "Abbr":     abbr,
                "Governor": gov,
                "Chambers": chamber_str,
                "Trifecta": trifecta,
                "Market":   market,
                "Bill":     bill_no,
                "Summary":  text,
                "URL":      url,
            })
    return rows


def write_excel(rows: list[dict], out: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bills"

    headers = ["State", "Abbr", "Governor", "Chambers", "Trifecta", "Market", "Bill", "Summary", "URL"]

    # Header style
    hdr_fill = PatternFill("solid", fgColor="1A2E4A")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    thin     = Side(style="thin", color="CCCCCC")
    border   = Border(bottom=Side(style="thin", color="AAAAAA"))

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font    = hdr_font
        cell.fill    = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 22

    # Party fill colours
    R_FILL = PatternFill("solid", fgColor="FDE8E8")
    D_FILL = PatternFill("solid", fgColor="E8EEF5")
    M_FILLS = {
        "Regulated":   PatternFill("solid", fgColor="E8F0E8"),
        "Deregulated": PatternFill("solid", fgColor="E8EEF5"),
        "Partial":     PatternFill("solid", fgColor="F5F0E0"),
    }

    for row_idx, row in enumerate(rows, 2):
        gov = row["Governor"]
        for col, key in enumerate(headers, 1):
            val  = row[key]
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font      = Font(size=9)
            cell.alignment = Alignment(vertical="top", wrap_text=(key == "Summary"))
            cell.border    = Border(bottom=Side(style="thin", color="EEEEEE"))

            # Colour coding
            if key == "Governor":
                cell.fill = R_FILL if gov == "R" else (D_FILL if gov == "D" else PatternFill())
            if key == "Trifecta":
                t = row["Trifecta"]
                cell.fill = R_FILL if t == "R-trifecta" else (D_FILL if t == "D-trifecta" else PatternFill())
            if key == "Market":
                cell.fill = M_FILLS.get(row["Market"], PatternFill())
            if key == "URL" and val:
                cell.hyperlink = val
                cell.font = Font(size=9, color="1A6AAA", underline="single")
                cell.value = "Link"

    # Column widths
    widths = {"State": 16, "Abbr": 6, "Governor": 9, "Chambers": 24,
              "Trifecta": 13, "Market": 13, "Bill": 14, "Summary": 70, "URL": 8}
    for col, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col)].width = widths.get(h, 12)

    # Freeze header row + enable autofilter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    wb.save(out)
    print(f"Saved {out}  ({len(rows)} rows)")


if __name__ == "__main__":
    data = load_states()
    rows = flatten(data)
    write_excel(rows, OUT_FILE)
