"""
Data center bill success-rate analysis by state party control.

Reads from the legiscan SQLite DB, classifies each bill via Groq (policy
direction + omnibus flag), joins with May 2026 party-control data, computes
passage outcomes, and writes a clean Excel workbook for visualization.

Run: python analyze_dc_bills.py
Output: dc_bills_analysis.xlsx

Party control source: NCSL + Ballotpedia, as of May 2026.
"""

import json
import os
import sys
import time
from pathlib import Path

from dotenv import load_dotenv
load_dotenv()

sys.path.insert(0, str(Path(__file__).parent))

from groq import Groq
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from legiscan.db import get_conn

OUT_FILE       = Path("dc_bills_analysis.xlsx")
CLASSIFY_CACHE = Path(".dc_classify_cache.json")

# ---------------------------------------------------------------------------
# Party control — NCSL + Ballotpedia, as of May 2026
# Format: abbr -> (governor_party, senate_party, house_party)
# Senate/House: "R", "D", "Split", "Nonpartisan"
# ---------------------------------------------------------------------------
PARTY_CONTROL: dict[str, tuple[str, str, str]] = {
    "AL": ("R", "R",            "R"),
    "AK": ("R", "R",            "R"),
    "AZ": ("D", "R",            "D"),
    "AR": ("R", "R",            "R"),
    "CA": ("D", "D",            "D"),
    "CO": ("D", "D",            "D"),
    "CT": ("D", "D",            "D"),
    "DE": ("D", "D",            "D"),
    "FL": ("R", "R",            "R"),
    "GA": ("R", "R",            "R"),
    "HI": ("D", "D",            "D"),
    "ID": ("R", "R",            "R"),
    "IL": ("D", "D",            "D"),
    "IN": ("R", "R",            "R"),
    "IA": ("R", "R",            "R"),
    "KS": ("D", "R",            "R"),
    "KY": ("D", "R",            "R"),
    "LA": ("R", "R",            "R"),
    "ME": ("D", "Split",        "Split"),
    "MD": ("D", "D",            "D"),
    "MA": ("D", "D",            "D"),
    "MI": ("D", "Split",        "Split"),
    "MN": ("D", "Split",        "Split"),
    "MS": ("R", "R",            "R"),
    "MO": ("R", "R",            "R"),
    "MT": ("R", "R",            "R"),
    "NE": ("R", "Nonpartisan",  "Nonpartisan"),  # unicameral, officially nonpartisan
    "NV": ("R", "D",            "D"),
    "NH": ("R", "R",            "R"),
    "NJ": ("D", "D",            "D"),
    "NM": ("D", "D",            "D"),
    "NY": ("D", "D",            "D"),
    "NC": ("D", "R",            "R"),
    "ND": ("R", "R",            "R"),
    "OH": ("R", "R",            "R"),
    "OK": ("R", "R",            "R"),
    "OR": ("D", "D",            "D"),
    "PA": ("D", "Split",        "Split"),
    "RI": ("D", "D",            "D"),
    "SC": ("R", "R",            "R"),
    "SD": ("R", "R",            "R"),
    "TN": ("R", "R",            "R"),
    "TX": ("R", "R",            "R"),
    "UT": ("R", "R",            "R"),
    "VT": ("R", "Split",        "D"),
    "VA": ("D", "D",            "D"),
    "WA": ("D", "D",            "D"),
    "WV": ("R", "R",            "R"),
    "WI": ("D", "R",            "Split"),
    "WY": ("R", "R",            "R"),
}

# LegiScan status_id → outcome bucket
OUTCOME_MAP = {
    0:  "Unknown",
    1:  "Active",   # Introduced
    2:  "Active",   # Engrossed
    3:  "Active",   # Enrolled (passed both chambers, awaiting signature)
    4:  "Passed",   # Passed
    5:  "Failed",   # Vetoed
    6:  "Failed",   # Failed
    7:  "Passed",   # Veto overridden
    8:  "Passed",   # Chaptered / signed into law
    9:  "Active",   # Referred
    10: "Active",   # Report Pass
    11: "Failed",   # Report DNP (Do Not Pass)
    12: "Active",   # Draft
}


def trifecta_label(gov: str, senate: str, house: str) -> str:
    if senate == "Nonpartisan":
        return "Divided (R Gov / Nonpartisan Leg)" if gov == "R" else "Divided (D Gov / Nonpartisan Leg)"
    both_chambers_r = senate == "R" and house == "R"
    both_chambers_d = senate == "D" and house == "D"
    if gov == "R" and both_chambers_r:
        return "R-Trifecta"
    if gov == "D" and both_chambers_d:
        return "D-Trifecta"
    return f"Divided ({gov} Gov)"


# ---------------------------------------------------------------------------
# LLM classification
# ---------------------------------------------------------------------------

CLASSIFY_PROMPT = """\
You are classifying state energy legislation related to data centers.

Bill: {state} {bill_number}
Title: {title}
Summary: {summary}
Tags: {tags}

Classify this bill on THREE dimensions:

1. policy_direction
   - "pro": supports, incentivizes, permits, or facilitates data center development,
     energy supply expansion, transmission buildout, or grid infrastructure for load growth.
     IMPORTANT: Tax exemptions, tax credits, and tax abatements FOR data centers = "pro"
     (they are government incentives, not burdens). Bills creating a legal "Data Center
     Definition" to enable future incentives = "pro". Bills providing data center funding,
     grants, or siting support = "pro".
   - "restrictive": places limits, fees, taxes, disclosure requirements, moratoriums,
     or siting/permitting burdens on data centers or energy-intensive industries.
     IMPORTANT: A new "Data Center Tax" or fee imposed ON data centers = "restrictive".
     Moratoriums on data center construction = "restrictive".
   - "neutral": study commissions, monitoring, reporting requirements, or process
     bills with no clear pro-development or anti-development stance.
     When evidence points toward pro or restrictive, commit to that — do not default
     to neutral out of uncertainty.

2. is_omnibus: true if this is a budget, appropriations, or omnibus bill that
   covers many unrelated policy areas; false otherwise

3. key_mechanism: the primary policy mechanism used. Pick the best match from this list:
   Tax Exemption, Moratorium, Utility Rates, Energy Planning, Zoning/Land Use,
   Transparency, Water, Data Center Tax, Data Center Funding, Employment,
   Air Quality, Renewables, Transmission, National Security, Data Center Definition,
   Noise, Permitting, Grid Reliability, Other

Return ONLY valid JSON: {{"policy_direction": "pro"|"restrictive"|"neutral", "is_omnibus": true|false, "key_mechanism": "..."}}"""


def load_cache() -> dict:
    if CLASSIFY_CACHE.exists():
        try:
            return json.loads(CLASSIFY_CACHE.read_text(encoding="utf-8"))
        except Exception:
            return {}
    return {}


def save_cache(cache: dict):
    CLASSIFY_CACHE.write_text(json.dumps(cache, indent=2, ensure_ascii=False), encoding="utf-8")


def classify_bill(client: Groq, bill_id: int, state: str, bill_number: str,
                  title: str, summary: str, tags: str, cache: dict) -> dict:
    key = str(bill_id)
    if key in cache and "key_mechanism" in cache[key]:
        return cache[key]

    prompt = CLASSIFY_PROMPT.format(
        state=state, bill_number=bill_number,
        title=title, summary=(summary or title or "")[:600],
        tags=tags or "",
    )

    for attempt in range(4):
        try:
            resp = client.chat.completions.create(
                model="llama-3.1-8b-instant",  # 500k TPD limit vs 100k for 70b
                messages=[{"role": "user", "content": prompt}],
                max_tokens=80,
                temperature=0,
                response_format={"type": "json_object"},
            )
            raw = resp.choices[0].message.content.strip()
            if not raw:
                raise ValueError("empty response")
            parsed = json.loads(raw)
            result = {
                "policy_direction": parsed.get("policy_direction", "neutral"),
                "is_omnibus":       bool(parsed.get("is_omnibus", False)),
                "key_mechanism":    parsed.get("key_mechanism", ""),
            }
            cache[key] = result
            save_cache(cache)
            return result
        except Exception as e:
            err = str(e)
            is_tpm = "tokens per minute" in err or "TPM" in err
            is_tpd = "tokens per day" in err or "TPD" in err
            is_503 = "503" in err or "over capacity" in err.lower()

            if is_tpd:
                # Daily limit hit — no point retrying, but don't cache failure
                print(f"    TPD limit hit, stopping classification", flush=True)
                return {"policy_direction": "neutral", "is_omnibus": False, "key_mechanism": ""}
            elif is_tpm or is_503:
                # Per-minute limit — wait and retry
                wait = 12 * (attempt + 1)
                print(f"    TPM limit, waiting {wait}s ...", flush=True)
                time.sleep(wait)
            elif attempt < 3:
                time.sleep(3)
            else:
                print(f"    classify failed {state} {bill_number}: {e}", flush=True)
                result = {"policy_direction": "neutral", "is_omnibus": False, "key_mechanism": ""}
                cache[key] = result
                save_cache(cache)
                return result

    return {"policy_direction": "neutral", "is_omnibus": False, "key_mechanism": ""}


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------

def load_bills() -> list[dict]:
    with get_conn() as conn:
        rows = conn.execute("""
            SELECT
                q.id, q.bill_id, q.state, q.bill_number, q.title,
                q.summary, q.tags, q.confidence, q.url, q.review_status,
                q.created_at,
                COALESCE(b.status_id, 0) AS status_id,
                b.last_action
            FROM queue q
            LEFT JOIN bills b ON q.bill_id = b.bill_id
            WHERE q.review_status != 'rejected'
            ORDER BY q.state, q.bill_number
        """).fetchall()
    return [dict(r) for r in rows]


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

R_FILL  = PatternFill("solid", fgColor="FDEAEA")
D_FILL  = PatternFill("solid", fgColor="E8EEF8")
NO_FILL = PatternFill()

OUTCOME_FILLS = {
    "Passed": PatternFill("solid", fgColor="D6F0D6"),
    "Failed": PatternFill("solid", fgColor="F5D5D5"),
    "Active": PatternFill("solid", fgColor="FFF8DC"),
    "Unknown": NO_FILL,
}

DIRECTION_FILLS = {
    "pro":        PatternFill("solid", fgColor="E8F5E9"),
    "restrictive": PatternFill("solid", fgColor="FDEAEA"),
    "neutral":    PatternFill("solid", fgColor="F5F5F5"),
}

TRIFECTA_FILLS = {
    "R-Trifecta":  PatternFill("solid", fgColor="FDEAEA"),
    "D-Trifecta":  PatternFill("solid", fgColor="E8EEF8"),
}

HDR_FILL = PatternFill("solid", fgColor="1A2E4A")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)


def write_header(ws, headers: list[str], widths: dict[str, int]):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = HDR_FONT
        cell.fill      = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22
    for col, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col)].width = widths.get(h, 12)
    ws.freeze_panes = f"A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def write_bills_sheet(ws, bills: list[dict]):
    headers = [
        "State", "Bill", "Title", "Outcome", "Policy Direction", "Key Mechanism", "Omnibus?",
        "Trifecta", "Governor", "Senate", "House",
        "Confidence", "Review Status", "Tags", "Summary", "URL",
    ]
    widths = {
        "State": 8, "Bill": 10, "Title": 40, "Outcome": 10,
        "Policy Direction": 15, "Key Mechanism": 22, "Omnibus?": 10,
        "Trifecta": 26, "Governor": 10, "Senate": 10, "House": 10,
        "Confidence": 12, "Review Status": 14, "Tags": 28,
        "Summary": 70, "URL": 8,
    }
    write_header(ws, headers, widths)

    for row_i, b in enumerate(bills, 2):
        vals = {
            "State":           b["state"],
            "Bill":            b["bill_number"],
            "Title":           b["title"],
            "Outcome":         b["outcome"],
            "Policy Direction": b["policy_direction"],
            "Key Mechanism":   b["key_mechanism"],
            "Omnibus?":        "Yes" if b["is_omnibus"] else "No",
            "Trifecta":        b["trifecta"],
            "Governor":        b["gov"],
            "Senate":          b["senate"],
            "House":           b["house"],
            "Confidence":      round(b["confidence"] or 0, 2),
            "Review Status":   b["review_status"],
            "Tags":            b["tags"],
            "Summary":         b["summary"],
            "URL":             b["url"],
        }
        for col, h in enumerate(headers, 1):
            val  = vals[h]
            cell = ws.cell(row=row_i, column=col, value=val)
            cell.font      = Font(size=9)
            cell.alignment = Alignment(vertical="top", wrap_text=(h in ("Summary", "Title")))
            cell.border    = Border(bottom=Side(style="thin", color="EEEEEE"))

            if h == "Outcome":
                cell.fill = OUTCOME_FILLS.get(val, NO_FILL)
            elif h == "Policy Direction":
                cell.fill = DIRECTION_FILLS.get(val, NO_FILL)
            elif h in ("Governor", "Senate", "House"):
                cell.fill = R_FILL if val == "R" else (D_FILL if val == "D" else NO_FILL)
            elif h == "Trifecta":
                cell.fill = TRIFECTA_FILLS.get(val, NO_FILL)
            elif h == "URL" and val:
                cell.hyperlink = val
                cell.font      = Font(size=9, color="1A6AAA", underline="single")
                cell.value     = "Link"


def write_summary_sheet(ws, bills: list[dict]):
    headers = [
        "Trifecta", "Policy Direction",
        "Total Bills", "Passed", "Failed", "Active", "Unknown",
        "Pass Rate (excl. Active/Unknown)",
        "Pass Rate (of Total)",
    ]
    widths = {
        "Trifecta": 28, "Policy Direction": 17,
        "Total Bills": 12, "Passed": 9, "Failed": 9, "Active": 9, "Unknown": 9,
        "Pass Rate (excl. Active/Unknown)": 30,
        "Pass Rate (of Total)": 22,
    }
    write_header(ws, headers, widths)

    # Aggregate
    from collections import defaultdict
    buckets: dict[tuple, dict] = defaultdict(lambda: {"Passed": 0, "Failed": 0, "Active": 0, "Unknown": 0})
    for b in bills:
        if b["is_omnibus"]:
            continue
        key = (b["trifecta"], b["policy_direction"])
        buckets[key][b["outcome"]] += 1

    row_i = 2
    trifecta_order = ["R-Trifecta", "D-Trifecta", "Divided (R Gov)", "Divided (D Gov)",
                      "Divided (R Gov / Nonpartisan Leg)", "Divided (D Gov / Nonpartisan Leg)"]
    direction_order = ["pro", "restrictive", "neutral"]

    seen_keys = set(buckets.keys())
    ordered_keys = [
        (t, d) for t in trifecta_order for d in direction_order
        if (t, d) in seen_keys
    ]
    # add any remaining (unexpected) keys
    for k in seen_keys:
        if k not in ordered_keys:
            ordered_keys.append(k)

    for (trifecta, direction) in ordered_keys:
        counts = buckets[(trifecta, direction)]
        total   = sum(counts.values())
        passed  = counts["Passed"]
        failed  = counts["Failed"]
        active  = counts["Active"]
        unknown = counts["Unknown"]
        decided = passed + failed
        rate_decided = (passed / decided) if decided > 0 else None
        rate_total   = (passed / total)   if total > 0   else None

        row = [
            trifecta, direction,
            total, passed, failed, active, unknown,
            f"{rate_decided:.0%}" if rate_decided is not None else "—",
            f"{rate_total:.0%}"   if rate_total   is not None else "—",
        ]
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=row_i, column=col, value=val)
            cell.font      = Font(size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = Border(bottom=Side(style="thin", color="EEEEEE"))
            h = headers[col - 1]
            if h == "Trifecta":
                cell.fill = TRIFECTA_FILLS.get(trifecta, NO_FILL)
                cell.alignment = Alignment(horizontal="left")
            elif h == "Policy Direction":
                cell.fill = DIRECTION_FILLS.get(direction, NO_FILL)
                cell.alignment = Alignment(horizontal="left")
            elif h in ("Pass Rate (excl. Active/Unknown)", "Pass Rate (of Total)"):
                cell.font = Font(size=9, bold=True)
        row_i += 1

    # Totals row
    row_i += 1
    ws.cell(row=row_i, column=1, value="TOTAL (excl. omnibus)").font = Font(bold=True, size=9)
    all_non_omnibus = [b for b in bills if not b["is_omnibus"]]
    for col, outcome in enumerate(["Passed", "Failed", "Active", "Unknown"], 4):
        ws.cell(row=row_i, column=col, value=sum(1 for b in all_non_omnibus if b["outcome"] == outcome)).font = Font(bold=True, size=9)


def write_party_control_sheet(ws):
    headers = ["State", "Abbr", "Governor", "Senate", "House", "Trifecta"]
    widths  = {"State": 18, "Abbr": 7, "Governor": 10, "Senate": 10, "House": 10, "Trifecta": 28}

    STATE_NAMES = {
        "AL": "Alabama", "AK": "Alaska", "AZ": "Arizona", "AR": "Arkansas",
        "CA": "California", "CO": "Colorado", "CT": "Connecticut", "DE": "Delaware",
        "FL": "Florida", "GA": "Georgia", "HI": "Hawaii", "ID": "Idaho",
        "IL": "Illinois", "IN": "Indiana", "IA": "Iowa", "KS": "Kansas",
        "KY": "Kentucky", "LA": "Louisiana", "ME": "Maine", "MD": "Maryland",
        "MA": "Massachusetts", "MI": "Michigan", "MN": "Minnesota", "MS": "Mississippi",
        "MO": "Missouri", "MT": "Montana", "NE": "Nebraska", "NV": "Nevada",
        "NH": "New Hampshire", "NJ": "New Jersey", "NM": "New Mexico", "NY": "New York",
        "NC": "North Carolina", "ND": "North Dakota", "OH": "Ohio", "OK": "Oklahoma",
        "OR": "Oregon", "PA": "Pennsylvania", "RI": "Rhode Island", "SC": "South Carolina",
        "SD": "South Dakota", "TN": "Tennessee", "TX": "Texas", "UT": "Utah",
        "VT": "Vermont", "VA": "Virginia", "WA": "Washington", "WV": "West Virginia",
        "WI": "Wisconsin", "WY": "Wyoming",
    }

    write_header(ws, headers, widths)

    # Note row
    note_cell = ws.cell(row=2, column=1, value="Party control as of May 2026  |  Source: NCSL + Ballotpedia")
    note_cell.font = Font(italic=True, size=9, color="666666")
    ws.merge_cells(f"A2:{get_column_letter(len(headers))}2")

    for row_i, (abbr, (gov, senate, house)) in enumerate(sorted(PARTY_CONTROL.items()), 3):
        tri = trifecta_label(gov, senate, house)
        vals = [STATE_NAMES.get(abbr, abbr), abbr, gov, senate, house, tri]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row_i, column=col, value=val)
            cell.font      = Font(size=9)
            cell.alignment = Alignment(vertical="center")
            cell.border    = Border(bottom=Side(style="thin", color="EEEEEE"))
            h = headers[col - 1]
            if h in ("Governor", "Senate", "House"):
                cell.fill = R_FILL if val == "R" else (D_FILL if val == "D" else NO_FILL)
            elif h == "Trifecta":
                cell.fill = TRIFECTA_FILLS.get(tri, NO_FILL)


# ---------------------------------------------------------------------------
# Methodology sheet
# ---------------------------------------------------------------------------

METHODOLOGY = [
    ("SECTION", "DETAIL"),  # header row
    ("", ""),
    ("OVERVIEW", ""),
    ("Purpose",
     "Analyze data center energy legislation across all 50 US states. "
     "Measure how often pro-development vs. restrictive bills pass under different "
     "state political configurations (R-Trifecta, D-Trifecta, Divided government)."),
    ("Research question",
     "Do states with unified Republican control pass more pro-data-center bills? "
     "Do states with unified Democratic control pass more restrictive bills? "
     "What is the baseline passage rate for each party configuration?"),
    ("", ""),
    ("BILL DATA", ""),
    ("Source", "LegiScan REST API (https://legiscan.com) — commercial legislative tracking service"),
    ("API calls used", "getSearch (full-text keyword search), getBill (bill metadata + status), getBillText (bill text)"),
    ("Search queries",
     "\"data center load\", \"data center power\", \"datacenter\", \"colocation facility\""),
    ("Session scope", "Current + prior legislative session (LegiScan year param = 3)"),
    ("Geographic scope", "All 50 US states"),
    ("Negative filter",
     "Bills containing signals clearly unrelated to energy (e.g., agriculture-only terms) "
     "are excluded before queuing. Implemented in legiscan/filter.py."),
    ("Fetch date", "May-June 2026"),
    ("", ""),
    ("PARTY CONTROL DATA", ""),
    ("Source", "NCSL (National Conference of State Legislatures) + Ballotpedia"),
    ("As-of date",
     "May 2026  —  NOTE: This is a snapshot. Bills introduced in prior sessions "
     "may have been introduced under a different party configuration."),
    ("Governor data", "Ballotpedia list of current governors"),
    ("Legislature data", "NCSL partisan composition page (https://www.ncsl.org/about-state-legislatures/state-partisan-composition)"),
    ("Nebraska note",
     "Nebraska has a unicameral, officially nonpartisan legislature. "
     "It is categorized separately as 'Divided (R Gov / Nonpartisan Leg)'."),
    ("", ""),
    ("TRIFECTA DEFINITIONS", ""),
    ("R-Trifecta", "Republican governor + Republican majority in both legislative chambers"),
    ("D-Trifecta", "Democratic governor + Democratic majority in both legislative chambers"),
    ("Divided (R Gov)",
     "Republican governor with at least one Democratic chamber or split chamber"),
    ("Divided (D Gov)",
     "Democratic governor with at least one Republican chamber or split chamber"),
    ("Split chamber", "Neither party holds a majority in that chamber"),
    ("", ""),
    ("BILL OUTCOME CLASSIFICATION", ""),
    ("Source field", "LegiScan status_id field from getBill response"),
    ("Passed",
     "status_id 4 (Passed), 7 (Veto overridden), 8 (Chaptered / signed into law)"),
    ("Failed",
     "status_id 5 (Vetoed), 6 (Failed), 11 (Report DNP — Do Not Pass)"),
    ("Active / Pending",
     "status_id 1 (Introduced), 2 (Engrossed), 3 (Enrolled), 9 (Referred), "
     "10 (Report Pass), 12 (Draft)  —  still moving through the process"),
    ("Unknown", "status_id 0 — no outcome data returned by API"),
    ("Pass rate formula (Summary sheet)",
     "Pass Rate (excl. Active/Unknown) = Passed ÷ (Passed + Failed). "
     "Bills still active are excluded so the rate reflects decided bills only. "
     "Pass Rate (of Total) = Passed ÷ All bills including Active."),
    ("", ""),
    ("POLICY DIRECTION CLASSIFICATION", ""),
    ("Model", "Groq llama-3.1-8b-instant (500k tokens/day free tier)"),
    ("Classification prompt",
     "Bills are classified as: "
     "PRO (supports/incentivizes data center development, energy supply, transmission, grid infrastructure), "
     "RESTRICTIVE (limits, fees, taxes, moratoriums, disclosure mandates on data centers or heavy energy users), "
     "NEUTRAL (study commissions, monitoring, reporting — no clear development stance)"),
    ("Input to model", "State, bill number, title, summary/description (up to 600 chars), keyword tags"),
    ("Output format", "JSON with policy_direction and is_omnibus fields, temperature=0"),
    ("Caching", "Results cached in .dc_classify_cache.json — bills are not re-classified on re-runs"),
    ("", ""),
    ("OMNIBUS / BUDGET BILL FILTER", ""),
    ("Definition",
     "Bills classified as omnibus=true are budget, appropriations, or catch-all bills "
     "that cover many unrelated policy areas. Their passage rate is artificially high "
     "and would distort the analysis."),
    ("Treatment",
     "Omnibus bills appear in the Bills sheet with 'Yes' in the Omnibus? column "
     "but are EXCLUDED from pass-rate calculations in the Summary sheet."),
    ("", ""),
    ("KNOWN LIMITATIONS", ""),
    ("Party control timing",
     "Party control is recorded as of May 2026, not at the time each bill was introduced. "
     "Bills introduced in 2023 under a different configuration will be miscategorized "
     "if the state changed hands after an election."),
    ("Active bills",
     "Bills still in 'Active' status may eventually pass or fail — they are not yet decided. "
     "Summary rates calculated excluding Active bills are more reliable."),
    ("Search coverage",
     "LegiScan full-text search may miss bills where 'data center' language appears "
     "only in committee amendments not indexed in the API. Recall is high but not 100%."),
    ("Classification accuracy",
     "LLM policy direction classification uses title + short description only "
     "for bills where full bill text was unavailable (common for GA PDF-only bills). "
     "Titles can be misleading — treat classifications as indicative, not definitive."),
    ("Small cell sizes",
     "Some trifecta × policy-direction buckets may have very few bills. "
     "Pass rates in cells with N < 10 should be interpreted with caution."),
    ("", ""),
    ("TECHNICAL PIPELINE", ""),
    ("Fetch script", "fetch_national.py — LegiScan API → legiscan.db (SQLite)"),
    ("Analysis script", "analyze_dc_bills.py — DB + Groq classification → dc_bills_analysis.xlsx"),
    ("Review tool", "legiscan/review_server.py — manual curation UI at localhost:8765"),
    ("DB location", "legiscan.db in project root (bills, queue, docs, watched_bills tables)"),
]


def write_methodology_sheet(ws):
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 90

    title_font   = Font(bold=True, size=14, color="1A2E4A")
    section_font = Font(bold=True, size=10, color="FFFFFF")
    section_fill = PatternFill("solid", fgColor="1A2E4A")
    key_font     = Font(bold=True, size=9)
    val_font     = Font(size=9)
    hdr_fill     = PatternFill("solid", fgColor="E8EEF8")

    # Title
    ws.merge_cells("A1:B1")
    title_cell = ws.cell(row=1, column=1,
        value="Data Center Energy Legislation Analysis — Methodology & Data Sources")
    title_cell.font      = title_font
    title_cell.fill      = hdr_fill
    title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:B2")
    sub = ws.cell(row=2, column=1, value="Generated by analyze_dc_bills.py  |  Party control data as of May 2026")
    sub.font      = Font(italic=True, size=9, color="666666")
    sub.alignment = Alignment(indent=1)
    ws.row_dimensions[2].height = 16

    row_i = 3
    for key, val in METHODOLOGY:
        if key == "SECTION":
            continue  # skip sentinel header

        if not key and not val:
            ws.row_dimensions[row_i].height = 6
            row_i += 1
            continue

        # All-caps key with no value → section header
        if key and key == key.upper() and not val:
            ws.merge_cells(f"A{row_i}:B{row_i}")
            cell = ws.cell(row=row_i, column=1, value=key)
            cell.font      = section_font
            cell.fill      = section_fill
            cell.alignment = Alignment(indent=1, vertical="center")
            ws.row_dimensions[row_i].height = 18
            row_i += 1
            continue

        # Normal key-value row
        key_cell = ws.cell(row=row_i, column=1, value=key)
        key_cell.font      = key_font
        key_cell.alignment = Alignment(vertical="top", wrap_text=True, indent=1)
        key_cell.border    = Border(bottom=Side(style="thin", color="EEEEEE"))

        val_cell = ws.cell(row=row_i, column=2, value=val)
        val_cell.font      = val_font
        val_cell.alignment = Alignment(vertical="top", wrap_text=True, indent=1)
        val_cell.border    = Border(bottom=Side(style="thin", color="EEEEEE"))
        ws.row_dimensions[row_i].height = 30 if len(val) > 80 else 16
        row_i += 1


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("Loading bills from DB ...", flush=True)
    bills = load_bills()
    print(f"  {len(bills)} bills (pending + approved)", flush=True)

    if not bills:
        print("No bills in queue yet — run fetch_national.py first.")
        return

    groq_client = Groq(api_key=os.environ["GROQ_API_KEY"])
    cache = load_cache()

    # 6000 TPM limit / ~530 tokens per bill ≈ 11 bills/min → 1 every 6s to stay safe
    RATE_DELAY = 6.0

    print("Classifying bills via Groq ...", flush=True)
    for i, b in enumerate(bills, 1):
        needs_api = str(b["bill_id"]) not in cache
        cls = classify_bill(
            groq_client, b["bill_id"], b["state"], b["bill_number"],
            b["title"], b["summary"] or "", b["tags"] or "", cache,
        )
        b["policy_direction"] = cls["policy_direction"]
        b["is_omnibus"]       = cls["is_omnibus"]
        b["key_mechanism"]    = cls.get("key_mechanism", "")

        gov, senate, house = PARTY_CONTROL.get(b["state"], ("?", "?", "?"))
        b["gov"]     = gov
        b["senate"]  = senate
        b["house"]   = house
        b["trifecta"] = trifecta_label(gov, senate, house)
        b["outcome"]  = OUTCOME_MAP.get(b["status_id"], "Unknown")

        if needs_api:
            time.sleep(RATE_DELAY)
        if i % 10 == 0 or i == len(bills):
            print(f"  {i}/{len(bills)} classified", flush=True)

    print("Writing Excel ...", flush=True)
    wb = openpyxl.Workbook()

    ws_bills = wb.active
    ws_bills.title = "Bills"
    write_bills_sheet(ws_bills, bills)

    ws_summary = wb.create_sheet("Summary")
    write_summary_sheet(ws_summary, bills)

    ws_control = wb.create_sheet("Party Control (May 2026)")
    write_party_control_sheet(ws_control)

    ws_method = wb.create_sheet("Methodology")
    write_methodology_sheet(ws_method)

    wb.save(OUT_FILE)
    total   = len(bills)
    passed  = sum(1 for b in bills if b["outcome"] == "Passed")
    failed  = sum(1 for b in bills if b["outcome"] == "Failed")
    active  = sum(1 for b in bills if b["outcome"] == "Active")
    omnibus = sum(1 for b in bills if b["is_omnibus"])
    print(f"\nSaved {OUT_FILE}")
    print(f"  {total} bills  |  {passed} passed  {failed} failed  {active} active  |  {omnibus} omnibus (excluded from summary)")


if __name__ == "__main__":
    main()
